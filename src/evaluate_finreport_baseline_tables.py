from __future__ import annotations

import argparse
import json
import math
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

try:
    from scipy.stats import f as f_dist
    from scipy.stats import chi2
    from scipy.stats import norm
except Exception:  # pragma: no cover
    f_dist = None
    chi2 = None
    norm = None


EPS = 1e-12


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Generate FinReport paper-style baseline and VaR evaluation tables.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--data-path", type=str, required=True, help="CSV file or folder containing train.csv/val.csv/test.csv")
    p.add_argument("--split", type=str, default="test", choices=["train", "val", "test", "all"], help="Which split to evaluate")
    p.add_argument("--output-dir", type=str, default="finreport_eval_outputs", help="Output folder")

    # Column config
    p.add_argument("--date-col", type=str, default=None, help="Date column. Auto-detect if omitted")
    p.add_argument("--code-col", type=str, default=None, help="Stock code column. Auto-detect if omitted")
    p.add_argument("--return-col", type=str, default=None, help="Realized return column. Auto-detect or compute from open/close if omitted")
    p.add_argument("--open-col", type=str, default="open1", help="Open column used to compute return when --return-col is omitted")
    p.add_argument("--close-col", type=str, default="close1", help="Close column used to compute return when --return-col is omitted")
    p.add_argument("--rf-col", type=str, default=None, help="Risk-free rate column. If omitted, RF=0")
    p.add_argument("--factor-cols", type=str, default=None, help="Comma-separated FF5 columns, e.g. MKT_RF,SMB,HML,RMW,CMA")
    p.add_argument("--news-col", type=str, default=None, help="News factor column. Auto-detect if omitted")
    p.add_argument("--var-col", type=str, default=None, help="Predicted VaR column. If omitted, rolling historical VaR is used")
    p.add_argument("--var-target-col", type=str, default=None, help="Actual VaR target column for RMSE/MAE. Optional")

    # Regression / GRS config
    p.add_argument("--allow-proxy-ff5", action="store_true", help="Create FF5-like proxy factors if real FF5 columns are not found")
    p.add_argument("--min-stock-obs", type=int, default=20, help="Minimum observations per stock for regression/risk stats")
    p.add_argument("--max-grs-stocks", type=int, default=100, help="Max number of stocks used in GRS covariance computation")
    p.add_argument("--seed", type=int, default=42, help="Random seed for stock subsampling when needed")

    # VaR config
    p.add_argument("--var-alpha", type=float, default=0.05, help="VaR left-tail alpha, e.g. 0.05 for 5% VaR")
    p.add_argument("--var-window", type=int, default=60, help="Rolling window for historical VaR fallback")
    p.add_argument("--annualization-days", type=int, default=252, help="Annualization factor for VaR")

    # Output config
    p.add_argument("--title", type=str, default="PAPER FINREPORT BASELINE", help="Table title")
    p.add_argument("--write-xlsx", action="store_true", default=True, help="Write formatted XLSX if openpyxl is available")
    p.add_argument("--no-xlsx", action="store_false", dest="write_xlsx", help="Do not write XLSX")
    p.add_argument("--percent-metrics", action="store_true", default=True, help="Write risk values as percentage-style values")
    return p.parse_args()


# ---------------------------------------------------------------------------
# Loading / preprocessing
# ---------------------------------------------------------------------------


def _read_csv_robust(path: Path) -> pd.DataFrame:
    """Read CSV/TSV robustly for the FinReport raw files."""
    # Many FinReport files are TSV-like. Try tab first, then comma.
    for sep in ["\t", ","]:
        try:
            df = pd.read_csv(path, sep=sep, low_memory=False, on_bad_lines="skip")
            if df.shape[1] > 1:
                return df
        except Exception:
            continue
    return pd.read_csv(path, low_memory=False, on_bad_lines="skip")


def load_data(data_path: str, split: str) -> pd.DataFrame:
    p = Path(data_path).expanduser().resolve()
    if not p.exists():
        raise FileNotFoundError(f"Data path not found: {p}")

    if p.is_file():
        return _read_csv_robust(p)

    if split == "all":
        frames = []
        for name in ["train.csv", "val.csv", "test.csv"]:
            fp = p / name
            if fp.exists():
                tmp = _read_csv_robust(fp)
                tmp["__split"] = name.replace(".csv", "")
                frames.append(tmp)
        if not frames:
            raise FileNotFoundError(f"No train.csv/val.csv/test.csv found in {p}")
        return pd.concat(frames, ignore_index=True)

    fp = p / f"{split}.csv"
    if not fp.exists():
        raise FileNotFoundError(f"Split file not found: {fp}")
    df = _read_csv_robust(fp)
    df["__split"] = split
    return df


def pick_column(df: pd.DataFrame, candidates: Sequence[str], name: str, required: bool = True) -> Optional[str]:
    lower_map = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand in df.columns:
            return cand
        if cand.lower() in lower_map:
            return lower_map[cand.lower()]
    if required:
        raise KeyError(f"Could not detect {name}. Tried: {candidates}")
    return None


def to_numeric_safe(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def prepare_base_dataframe(args: argparse.Namespace) -> Tuple[pd.DataFrame, Dict[str, str], List[str]]:
    notes: List[str] = []
    df = load_data(args.data_path, args.split)
    df = df.copy()

    date_col = args.date_col or pick_column(df, ["trade_date", "DATE", "date", "Date"], "date column")
    code_col = args.code_col or pick_column(df, ["CODE", "code", "ticker", "symbol", "stock", "Stock"], "stock code column")
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    df = df[df[date_col].notna()].copy()
    df["__date"] = df[date_col].dt.date.astype(str)
    df["__code"] = df[code_col].astype(str)

    # Realized return.
    ret_col = args.return_col
    if ret_col is None:
        ret_col = pick_column(
            df,
            ["CHANGE", "change", "return", "ret", "daily_return", "target_return", "future_return", "RET"],
            "return column",
            required=False,
        )
    if ret_col is not None and ret_col in df.columns:
        df["__ret"] = to_numeric_safe(df[ret_col])
        notes.append(f"Return column used: {ret_col}")
    else:
        if args.open_col not in df.columns or args.close_col not in df.columns:
            raise KeyError(
                f"No return column found and cannot compute from {args.open_col}/{args.close_col}. "
                "Pass --return-col explicitly."
            )
        o = to_numeric_safe(df[args.open_col])
        c = to_numeric_safe(df[args.close_col])
        df["__ret"] = (c - o) / (o.abs() + EPS)
        notes.append(f"Return computed as ({args.close_col}-{args.open_col})/{args.open_col}")

    # Risk-free rate.
    if args.rf_col and args.rf_col in df.columns:
        df["__rf"] = to_numeric_safe(df[args.rf_col]).fillna(0.0)
        notes.append(f"Risk-free column used: {args.rf_col}")
    else:
        df["__rf"] = 0.0
        notes.append("Risk-free rate not provided; RF=0 used.")

    df["__excess_ret"] = df["__ret"] - df["__rf"]
    df = df[np.isfinite(df["__excess_ret"])].copy()

    columns = {"date": date_col, "code": code_col, "return": ret_col or "computed", "rf": args.rf_col or "0"}
    return df, columns, notes


# ---------------------------------------------------------------------------
# FF5 factor construction / detection
# ---------------------------------------------------------------------------


def detect_factor_columns(df: pd.DataFrame, explicit: Optional[str]) -> Optional[List[str]]:
    if explicit:
        cols = [c.strip() for c in explicit.split(",") if c.strip()]
        missing = [c for c in cols if c not in df.columns]
        if missing:
            raise KeyError(f"Missing factor columns: {missing}")
        if len(cols) != 5:
            raise ValueError("--factor-cols must contain exactly 5 columns for FF5")
        return cols

    candidate_sets = [
        ["MKT_RF", "SMB", "HML", "RMW", "CMA"],
        ["mkt_rf", "smb", "hml", "rmw", "cma"],
        ["MKT", "SMB", "HML", "RMW", "CMA"],
        ["market", "size", "value", "profitability", "investment"],
        ["market_factor", "size_factor", "value_factor", "profitability_factor", "investment_factor"],
    ]
    lower_map = {c.lower(): c for c in df.columns}
    for cand_set in candidate_sets:
        out = []
        ok = True
        for cand in cand_set:
            if cand in df.columns:
                out.append(cand)
            elif cand.lower() in lower_map:
                out.append(lower_map[cand.lower()])
            else:
                ok = False
                break
        if ok:
            return out
    return None


def build_proxy_ff5_factors(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str], str]:
    """Build FF5-like daily proxy factors from returns and open/close-derived features.

    These are not true Fama-French factors. They are provided only when the raw
    dataset lacks official FF5 columns, so the user can still produce a diagnostic
    table with the same shape.
    """
    tmp = df.copy()

    # Proxy variables at stock-date level.
    if "open1" in tmp.columns:
        tmp["__size_proxy"] = to_numeric_safe(tmp["open1"])
    else:
        tmp["__size_proxy"] = tmp.groupby("__code")["__ret"].transform("std").fillna(0.0)

    # Momentum/reversal proxy from available open/close windows.
    if "open5" in tmp.columns and "close1" in tmp.columns:
        tmp["__mom_proxy"] = (to_numeric_safe(tmp["close1"]) - to_numeric_safe(tmp["open5"])) / (to_numeric_safe(tmp["open5"]).abs() + EPS)
    else:
        tmp["__mom_proxy"] = tmp.groupby("__code")["__ret"].transform(lambda s: s.shift(1).rolling(5, min_periods=1).mean())

    if "close1" in tmp.columns and "open1" in tmp.columns:
        tmp["__profit_proxy"] = (to_numeric_safe(tmp["close1"]) - to_numeric_safe(tmp["open1"])) / (to_numeric_safe(tmp["open1"]).abs() + EPS)
    else:
        tmp["__profit_proxy"] = tmp["__ret"]

    tmp["__vol_proxy"] = tmp.groupby("__code")["__ret"].transform(lambda s: s.shift(1).rolling(5, min_periods=2).std()).fillna(0.0)

    rows = []
    for d, g in tmp.groupby("__date"):
        ret = g["__ret"].astype(float)
        mkt = ret.mean()

        def high_low_return(score: pd.Series, high_minus_low: bool = True) -> float:
            score = pd.to_numeric(score, errors="coerce")
            valid = score.notna() & ret.notna()
            if valid.sum() < 4:
                return 0.0
            q_low = score[valid].quantile(0.3)
            q_high = score[valid].quantile(0.7)
            low_ret = ret[valid & (score <= q_low)].mean()
            high_ret = ret[valid & (score >= q_high)].mean()
            if not np.isfinite(low_ret) or not np.isfinite(high_ret):
                return 0.0
            return float(high_ret - low_ret if high_minus_low else low_ret - high_ret)

        smb = high_low_return(g["__size_proxy"], high_minus_low=False)       # small minus big
        hml = high_low_return(g["__mom_proxy"], high_minus_low=False)        # low momentum minus high momentum as value/reversal proxy
        rmw = high_low_return(g["__profit_proxy"], high_minus_low=True)      # robust-ish minus weak-ish
        cma = high_low_return(g["__vol_proxy"], high_minus_low=False)        # conservative-ish minus aggressive-ish via vol
        rows.append({"__date": d, "MKT_PROXY": mkt, "SMB_PROXY": smb, "HML_PROXY": hml, "RMW_PROXY": rmw, "CMA_PROXY": cma})

    fac = pd.DataFrame(rows)
    cols = ["MKT_PROXY", "SMB_PROXY", "HML_PROXY", "RMW_PROXY", "CMA_PROXY"]
    note = "Official FF5 columns not found; using FF5-like proxy factors derived from cross-sectional returns and price proxies."
    return fac, cols, note


def get_daily_factor_frame(df: pd.DataFrame, args: argparse.Namespace, notes: List[str]) -> Tuple[pd.DataFrame, List[str], bool]:
    factor_cols = detect_factor_columns(df, args.factor_cols)
    if factor_cols is not None:
        fac = df[["__date"] + factor_cols].copy()
        for c in factor_cols:
            fac[c] = to_numeric_safe(fac[c])
        fac = fac.groupby("__date", as_index=False)[factor_cols].mean()
        notes.append(f"Official/user-provided factor columns used: {factor_cols}")
        return fac, factor_cols, False

    if not args.allow_proxy_ff5:
        raise KeyError(
            "Could not find FF5 factor columns. Pass --factor-cols MKT_RF,SMB,HML,RMW,CMA "
            "or rerun with --allow-proxy-ff5 to create diagnostic proxy factors."
        )
    fac, cols, note = build_proxy_ff5_factors(df)
    notes.append(note)
    return fac, cols, True


def get_news_signal(df: pd.DataFrame, args: argparse.Namespace, notes: List[str]) -> pd.DataFrame:
    news_col = args.news_col
    if news_col is None:
        news_col = pick_column(
            df,
            ["news_factor", "news_score", "news_signal", "module1_score", "prob_UP", "prob_POSITIVE", "label"],
            "news factor column",
            required=False,
        )
    out = df[["__date", "__code"]].copy()
    if news_col and news_col in df.columns:
        out["NEWS_FACTOR"] = to_numeric_safe(df[news_col]).fillna(0.0)
        notes.append(f"News factor column used: {news_col}")
        if news_col.lower() == "label":
            notes.append("WARNING: news factor uses label; this can leak target information. Prefer out-of-sample Module I predictions.")
    else:
        out["NEWS_FACTOR"] = 0.0
        notes.append("News factor column not found; FF5+News uses NEWS_FACTOR=0. Provide --news-col for a meaningful FF5+News table.")
    return out


# ---------------------------------------------------------------------------
# Regression / GRS
# ---------------------------------------------------------------------------


def ols_fit(y: np.ndarray, X: np.ndarray) -> Tuple[np.ndarray, np.ndarray, float]:
    mask = np.isfinite(y) & np.all(np.isfinite(X), axis=1)
    y = y[mask].astype(float)
    X = X[mask].astype(float)
    if len(y) <= X.shape[1] + 2:
        raise ValueError("Insufficient observations for OLS")
    beta, *_ = np.linalg.lstsq(X, y, rcond=None)
    pred = X @ beta
    resid = y - pred
    sst = float(((y - y.mean()) ** 2).sum())
    sse = float((resid ** 2).sum())
    r2 = 1.0 - sse / (sst + EPS)
    return beta, resid, r2


@dataclass
class RegressionResult:
    model_name: str
    mean_abs_alpha: float
    mean_r2: float
    grs: float
    p_value: float
    n_stocks: int
    n_dates: int


def compute_regression_table(df: pd.DataFrame, factors: pd.DataFrame, factor_cols: List[str], news_df: pd.DataFrame, args: argparse.Namespace) -> Tuple[pd.DataFrame, Dict[str, object]]:
    base = df[["__date", "__code", "__excess_ret"]].merge(factors, on="__date", how="left")
    base = base.merge(news_df, on=["__date", "__code"], how="left")
    base["NEWS_FACTOR"] = base["NEWS_FACTOR"].fillna(0.0)
    for c in factor_cols:
        base[c] = to_numeric_safe(base[c])

    counts = base.groupby("__code").size().sort_values(ascending=False)
    eligible = counts[counts >= args.min_stock_obs].index.tolist()
    if len(eligible) == 0:
        raise RuntimeError("No stock has enough observations for regression. Lower --min-stock-obs.")

    # Use most-observed stocks for stable GRS.
    selected = eligible[: args.max_grs_stocks]
    rng = np.random.default_rng(args.seed)
    # If too many selected for available dates, later GRS may cap further.

    def run_model(model_name: str, xcols: List[str], common_factor_cols_for_grs: List[str]) -> Tuple[RegressionResult, pd.DataFrame, pd.DataFrame]:
        alphas, r2s = [], []
        resid_series = {}
        date_factors = None
        used_stocks = []

        for code in selected:
            g = base[base["__code"] == code].sort_values("__date")
            y = g["__excess_ret"].to_numpy(dtype=float)
            X_raw = g[xcols].to_numpy(dtype=float)
            X = np.column_stack([np.ones(len(g)), X_raw])
            try:
                beta, resid, r2 = ols_fit(y, X)
            except Exception:
                continue
            mask = np.isfinite(y) & np.all(np.isfinite(X), axis=1)
            valid_dates = g.loc[mask, "__date"].tolist()
            resid_series[code] = pd.Series(resid, index=valid_dates)
            alphas.append(float(beta[0]))
            r2s.append(float(r2))
            used_stocks.append(code)

        if not resid_series:
            raise RuntimeError(f"No valid regressions for {model_name}.")

        resid_df = pd.DataFrame(resid_series).sort_index()
        # Keep rows with enough observed residuals; fill remaining with 0 because OLS residuals are centered.
        min_non_na = max(5, int(0.6 * resid_df.shape[1]))
        resid_df = resid_df[resid_df.notna().sum(axis=1) >= min_non_na].fillna(0.0)

        # Cap N so T - N - K > 0.
        T = resid_df.shape[0]
        K = len(common_factor_cols_for_grs)
        max_n_valid = max(1, T - K - 2)
        if resid_df.shape[1] > max_n_valid:
            keep_cols = list(resid_df.columns[:max_n_valid])
            resid_df = resid_df[keep_cols]
            # Recompute alpha/r2 only among kept stocks if possible.
            # For simplicity, table metrics still use all valid single-stock regressions.

        common = factors.copy()
        if "NEWS_FACTOR_COMMON" in common_factor_cols_for_grs:
            news_common = news_df.groupby("__date", as_index=False)["NEWS_FACTOR"].mean().rename(columns={"NEWS_FACTOR": "NEWS_FACTOR_COMMON"})
            common = common.merge(news_common, on="__date", how="left")
        common = common.set_index("__date")
        common = common.reindex(resid_df.index)
        for c in common_factor_cols_for_grs:
            if c not in common.columns:
                common[c] = 0.0
        factor_mat = common[common_factor_cols_for_grs].astype(float).fillna(0.0).to_numpy()

        grs, pval = grs_test(np.asarray(alphas[: resid_df.shape[1]], dtype=float), resid_df.to_numpy(dtype=float), factor_mat)
        return (
            RegressionResult(
                model_name=model_name,
                mean_abs_alpha=float(np.nanmean(np.abs(alphas))),
                mean_r2=float(np.nanmean(r2s)),
                grs=float(grs) if np.isfinite(grs) else float("nan"),
                p_value=float(pval) if np.isfinite(pval) else float("nan"),
                n_stocks=int(len(used_stocks)),
                n_dates=int(T),
            ),
            resid_df,
            common.reset_index(),
        )

    res_ff5, _, _ = run_model("FF5", factor_cols, factor_cols)
    res_ff5_news, _, _ = run_model("FF5 + News", factor_cols + ["NEWS_FACTOR"], factor_cols + ["NEWS_FACTOR_COMMON"])

    rows = []
    for r in [res_ff5, res_ff5_news]:
        rows.append(
            {
                "Model": r.model_name,
                "Mean |α|": r.mean_abs_alpha,
                "Mean R²": r.mean_r2,
                "GRS": r.grs,
                "p-value": r.p_value,
                "N Stocks": r.n_stocks,
                "N Dates": r.n_dates,
            }
        )
    table = pd.DataFrame(rows)
    meta = {"selected_stocks_count": len(selected), "eligible_stocks_count": len(eligible), "factor_cols": factor_cols}
    return table, meta


def grs_test(alpha: np.ndarray, residuals: np.ndarray, factors: np.ndarray) -> Tuple[float, float]:
    """Compute approximate GRS test using pseudo-inverse for stability."""
    if f_dist is None:
        return float("nan"), float("nan")
    E = np.asarray(residuals, dtype=float)
    F = np.asarray(factors, dtype=float)
    alpha = np.asarray(alpha, dtype=float).reshape(-1, 1)
    T, N = E.shape
    if N <= 0 or T <= 3:
        return float("nan"), float("nan")
    K = F.shape[1] if F.ndim == 2 else 1
    if T - N - K <= 0:
        return float("nan"), float("nan")
    if alpha.shape[0] != N:
        # Align dimensions conservatively.
        m = min(alpha.shape[0], N)
        alpha = alpha[:m]
        E = E[:, :m]
        N = m
    sigma_e = np.cov(E, rowvar=False, bias=False)
    if N == 1:
        sigma_e = np.array([[float(np.var(E[:, 0], ddof=1))]])
    sigma_f = np.cov(F, rowvar=False, bias=False)
    if K == 1:
        sigma_f = np.array([[float(np.var(F[:, 0], ddof=1))]])
    mu_f = np.nanmean(F, axis=0).reshape(-1, 1)
    inv_sigma_e = np.linalg.pinv(sigma_e)
    inv_sigma_f = np.linalg.pinv(sigma_f)
    numerator = float(alpha.T @ inv_sigma_e @ alpha)
    denominator = float(1.0 + mu_f.T @ inv_sigma_f @ mu_f)
    grs = ((T - N - K) / max(N, 1)) * numerator / (denominator + EPS)
    p_value = 1.0 - float(f_dist.cdf(grs, N, T - N - K))
    return grs, p_value


# ---------------------------------------------------------------------------
# VaR / risk evaluation
# ---------------------------------------------------------------------------


def kupiec_pvalue(n: int, x: int, alpha: float) -> float:
    if chi2 is None or n <= 0:
        return float("nan")
    phat = min(max(x / n, EPS), 1 - EPS)
    a = min(max(alpha, EPS), 1 - EPS)
    # LR_uc = -2 ln(L0/L1)
    log_l0 = (n - x) * math.log(1 - a) + x * math.log(a)
    log_l1 = (n - x) * math.log(1 - phat) + x * math.log(phat)
    lr = -2.0 * (log_l0 - log_l1)
    return 1.0 - float(chi2.cdf(lr, df=1))


def build_var_predictions(df: pd.DataFrame, args: argparse.Namespace, notes: List[str]) -> pd.DataFrame:
    out = df[["__date", "__code", "__ret"]].copy()
    var_col = args.var_col
    if var_col is None:
        var_col = pick_column(df, ["var_est", "var_pred", "VaR", "VAR", "pred_var", "predicted_var"], "VaR column", required=False)
    if var_col and var_col in df.columns:
        out["var_pred"] = to_numeric_safe(df[var_col])
        notes.append(f"Predicted VaR column used: {var_col}")
    else:
        notes.append(f"Predicted VaR not provided; using rolling historical {int(args.var_alpha*100)}% VaR with window={args.var_window}.")
        out = out.sort_values(["__code", "__date"])
        out["var_pred"] = (
            out.groupby("__code")["__ret"]
            .transform(lambda s: s.shift(1).rolling(args.var_window, min_periods=max(5, args.var_window // 5)).quantile(args.var_alpha))
        )
        global_var = float(out["__ret"].quantile(args.var_alpha)) if len(out) else 0.0
        out["var_pred"] = out["var_pred"].fillna(global_var)

    if args.var_target_col and args.var_target_col in df.columns:
        out["var_target"] = to_numeric_safe(df[args.var_target_col])
        notes.append(f"VaR target column used for RMSE/MAE: {args.var_target_col}")
    else:
        out["var_target"] = np.nan
        notes.append("VaR target not provided; RMSE/MAE are computed against realized returns, not actual VaR targets.")
    out = out[np.isfinite(out["var_pred"]) & np.isfinite(out["__ret"])].copy()
    return out


def quantile_var_loss(returns: np.ndarray, var_pred: np.ndarray, alpha: float) -> float:
    """Pinball/quantile loss for VaR forecasts. Lower is better."""
    r = np.asarray(returns, dtype=float)
    v = np.asarray(var_pred, dtype=float)
    return float(np.mean(np.maximum(alpha * (r - v), (alpha - 1) * (r - v))))


def compute_risk_tables(df: pd.DataFrame, args: argparse.Namespace, notes: List[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    risk = build_var_predictions(df, args, notes)
    if risk.empty:
        raise RuntimeError("No valid rows for VaR evaluation")

    if risk["var_target"].notna().sum() > 0:
        err = risk["var_pred"] - risk["var_target"]
    else:
        err = risk["var_pred"] - risk["__ret"]

    rmse = float(np.sqrt(np.mean(np.square(err))))
    mae = float(np.mean(np.abs(err)))
    var_loss = quantile_var_loss(risk["__ret"].to_numpy(), risk["var_pred"].to_numpy(), args.var_alpha)

    risk["breach"] = risk["__ret"] < risk["var_pred"]
    day_coverage = 1.0 - float(risk["breach"].mean())

    # Stock-level stats.
    grouped = risk.groupby("__code").agg(
        n=("breach", "size"),
        breach_rate=("breach", "mean"),
        mean_var=("var_pred", "mean"),
    )
    grouped = grouped[grouped["n"] >= args.min_stock_obs].copy()
    if grouped.empty:
        grouped = risk.groupby("__code").agg(n=("breach", "size"), breach_rate=("breach", "mean"), mean_var=("var_pred", "mean"))

    kupiec_ps = []
    for code, g in risk.groupby("__code"):
        if len(g) < args.min_stock_obs:
            continue
        kupiec_ps.append(kupiec_pvalue(n=len(g), x=int(g["breach"].sum()), alpha=args.var_alpha))
    kupiec_ps = np.asarray([p for p in kupiec_ps if np.isfinite(p)], dtype=float)
    rejected_rate = float(np.mean(kupiec_ps < 0.05)) if len(kupiec_ps) else float("nan")

    mean_var = float(risk["var_pred"].mean())
    annualized_var = mean_var * math.sqrt(args.annualization_days)
    mean_breach_rate = float(grouped["breach_rate"].mean())
    median_breach_rate = float(grouped["breach_rate"].median())
    stock_level_coverage = float(np.mean((1.0 - grouped["breach_rate"]) >= (1.0 - args.var_alpha)))

    module3 = pd.DataFrame(
        [
            {"Metric": "RMSE", "Module 3 Result (%)": rmse * 100},
            {"Metric": "MAE", "Module 3 Result (%)": mae * 100},
            {"Metric": "VaR Loss", "Module 3 Result (%)": var_loss * 100},
            {"Metric": "Coverage Rate (day level)", "Module 3 Result (%)": day_coverage * 100},
        ]
    )
    risk_summary = pd.DataFrame(
        [
            {"Metric": "Stocks evaluated", "Value": int(grouped.shape[0])},
            {"Metric": f"Mean VaR ({int(args.var_alpha * 100)}%)", "Value": mean_var * 100},
            {"Metric": "Annualized VaR", "Value": annualized_var * 100},
            {"Metric": "Mean breach rate", "Value": mean_breach_rate * 100},
            {"Metric": "Median breach rate", "Value": median_breach_rate * 100},
            {"Metric": "Rejected by Kupiec test", "Value": rejected_rate * 100},
            {"Metric": "Coverage rate (day-level)", "Value": day_coverage * 100},
            {"Metric": "Coverage rate (stock-level)", "Value": stock_level_coverage * 100},
        ]
    )
    return module3, risk_summary


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------


def fmt_float(x: object, ndigits: int = 4) -> str:
    try:
        v = float(x)
        if not np.isfinite(v):
            return ""
        return f"{v:.{ndigits}f}"
    except Exception:
        return str(x)


def save_markdown_and_csv(ff_table: pd.DataFrame, mod3: pd.DataFrame, risk_summary: pd.DataFrame, out_dir: Path, notes: List[str]) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    ff_table.to_csv(out_dir / "ff5_baseline_table.csv", index=False, encoding="utf-8-sig")
    mod3.to_csv(out_dir / "module3_var_metrics.csv", index=False, encoding="utf-8-sig")
    risk_summary.to_csv(out_dir / "module3_var_summary.csv", index=False, encoding="utf-8-sig")

    md = []
    md.append("# Paper FinReport Baseline Tables\n")
    md.append("## FF5 Regression Summary\n")
    md.append(ff_table[["Model", "Mean |α|", "Mean R²", "GRS", "p-value"]].to_markdown(index=False))
    md.append("\n## Module 3 VaR Metrics\n")
    md.append(mod3.to_markdown(index=False))
    md.append("\n## VaR Summary\n")
    md.append(risk_summary.to_markdown(index=False))
    md.append("\n## Notes\n")
    for n in notes:
        md.append(f"- {n}")
    (out_dir / "finreport_baseline_tables.md").write_text("\n".join(md), encoding="utf-8")

    payload = {"notes": notes}
    (out_dir / "finreport_baseline_notes.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def save_xlsx(ff_table: pd.DataFrame, mod3: pd.DataFrame, risk_summary: pd.DataFrame, out_path: Path, title: str, notes: List[str]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception as e:
        warnings.warn(f"openpyxl not available; skipping XLSX output: {e}")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "FinReport Tables"

    green = "B6D7A8"
    light_green = "D9EAD3"
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def title_row(row: int, start_col: int, end_col: int, text: str):
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        c = ws.cell(row=row, column=start_col, value=text)
        c.font = Font(bold=False, size=16)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 28

    def write_table(row: int, col: int, headers: List[str], rows: List[List[object]], percent_cols: Optional[Iterable[int]] = None):
        percent_cols = set(percent_cols or [])
        for j, h in enumerate(headers):
            cell = ws.cell(row=row, column=col + j, value=h)
            cell.fill = PatternFill("solid", fgColor=green)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for i, values in enumerate(rows, start=1):
            for j, v in enumerate(values):
                cell = ws.cell(row=row + i, column=col + j, value=v)
                if j == 0:
                    cell.fill = PatternFill("solid", fgColor=light_green)
                if isinstance(v, (int, float, np.integer, np.floating)) and not isinstance(v, bool):
                    if j in percent_cols:
                        cell.number_format = "0.00%"
                        cell.value = float(v) / 100.0
                    elif abs(float(v)) < 10 and j != 0:
                        cell.number_format = "0.0000"
                    else:
                        cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center" if j > 0 else "left")
                cell.border = border

    # Table 1: FF5.
    title_row(1, 1, 5, title)
    ff_show = ff_table[["Model", "Mean |α|", "Mean R²", "GRS", "p-value"]].copy()
    ff_rows = ff_show.values.tolist()
    write_table(3, 1, list(ff_show.columns), ff_rows)

    # Table 2: Module 3 metrics.
    start = 8
    title_row(start, 1, 2, title)
    m3_rows = mod3.values.tolist()
    write_table(start + 2, 1, list(mod3.columns), m3_rows, percent_cols=[1])

    # Table 3: VaR Summary.
    start2 = start + 9
    risk_rows = risk_summary.values.tolist()
    write_table(start2, 1, list(risk_summary.columns), risk_rows, percent_cols=[1])
    # First row is count, not percent.
    count_cell = ws.cell(row=start2 + 1, column=2)
    count_cell.number_format = "0"
    # However previous formatting divided it by 100. Restore integer value.
    count_cell.value = int(risk_summary.iloc[0, 1])

    # Notes sheet.
    ns = wb.create_sheet("Notes")
    ns.cell(1, 1, "Notes / Data Caveats").font = Font(bold=True, size=14)
    for i, n in enumerate(notes, start=2):
        ns.cell(i, 1, n)
        ns.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ns.column_dimensions["A"].width = 120

    widths = {"A": 28, "B": 16, "C": 16, "D": 14, "E": 14}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "center",
                vertical="center",
                wrap_text=True,
            )
    wb.save(out_path)
    return True


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    args = parse_args()
    out_dir = Path(args.output_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    notes: List[str] = []
    df, columns, prep_notes = prepare_base_dataframe(args)
    notes.extend(prep_notes)
    factors, factor_cols, used_proxy = get_daily_factor_frame(df, args, notes)
    news_df = get_news_signal(df, args, notes)

    ff_table, regression_meta = compute_regression_table(df, factors, factor_cols, news_df, args)
    mod3, risk_summary = compute_risk_tables(df, args, notes)

    notes.append(f"Rows evaluated: {len(df):,}; split={args.split}.")
    notes.append(f"Regression metadata: {regression_meta}.")
    if used_proxy:
        notes.append("Interpret FF5 rows as diagnostic FF5-like proxy results, not official Fama-French replication.")

    save_markdown_and_csv(ff_table, mod3, risk_summary, out_dir, notes)
    if args.write_xlsx:
        xlsx_path = out_dir / "finreport_baseline_tables.xlsx"
        ok = save_xlsx(ff_table, mod3, risk_summary, xlsx_path, args.title, notes)
        if ok:
            print(f"[Saved] XLSX: {xlsx_path}")

    print(f"[Saved] CSV/MD/JSON outputs to: {out_dir}")
    print("\nFF5 Regression Summary")
    print(ff_table[["Model", "Mean |α|", "Mean R²", "GRS", "p-value"]].to_string(index=False))
    print("\nModule 3 VaR Metrics")
    print(mod3.to_string(index=False))
    print("\nVaR Summary")
    print(risk_summary.to_string(index=False))
    print("\nNotes:")
    for n in notes:
        print(f"- {n}")


if __name__ == "__main__":
    main()
